import os
import time
import shutil
import openpyxl
import pandas as pd
import numpy as np
import streamlit as st
import datetime
import io

# Page configuration
st.set_page_config(
    page_title="ISN Survey Data Integration Hub",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium styling
st.markdown("""
<style>
    /* Main container background */
    .reportview-container {
        background: #F8FAFC;
    }
    
    /* Header styling */
    .main-title {
        color: #1E3A8A;
        font-family: 'Outfit', 'Inter', sans-serif;
        font-weight: 800;
        font-size: 2.5rem;
        margin-bottom: 0.2rem;
    }
    .subtitle {
        color: #64748B;
        font-family: 'Inter', sans-serif;
        font-size: 1.1rem;
        margin-bottom: 2rem;
    }
    
    /* Card design */
    .metric-card {
        background-color: white;
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        border-left: 5px solid #0D9488;
        margin-bottom: 1rem;
    }
    .metric-card-spsc {
        border-left: 5px solid #3B82F6;
    }
    .metric-card-med {
        border-left: 5px solid #8B5CF6;
    }
    .card-title {
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        color: #64748B;
        margin-bottom: 0.5rem;
        font-weight: 700;
    }
    .card-value {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1E293B;
    }
    
    /* Section dividers */
    .section-header {
        color: #1E293B;
        font-family: 'Outfit', sans-serif;
        font-weight: 700;
        font-size: 1.4rem;
        margin-top: 2rem;
        margin-bottom: 1rem;
        border-bottom: 2px solid #E2E8F0;
        padding-bottom: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)

# Application state
if 'dry_run_completed' not in st.session_state:
    st.session_state.dry_run_completed = False
if 'spsc_piv_preview' not in st.session_state:
    st.session_state.spsc_piv_preview = None
if 'spsc_unpiv_preview' not in st.session_state:
    st.session_state.spsc_unpiv_preview = None
if 'med_piv_preview' not in st.session_state:
    st.session_state.med_piv_preview = None
if 'med_unpiv_preview' not in st.session_state:
    st.session_state.med_unpiv_preview = None
if 'raw_preview' not in st.session_state:
    st.session_state.raw_preview = None

# Check if running locally (default path exists)
default_master = r"c:\ISN CODING\ISN Survey\Data_for_AG\Survey 2022-2025 Data Source.xlsx"
default_raw = r"c:\ISN CODING\ISN Survey\Data_for_AG\2026 ISN PART 1_ Athletes and Coaches Satisfaction Survey on Sports Science & Sports Medicine Services (Responses).xlsx"
is_local_mode = os.path.exists(default_master)

# Sidebar - Settings and Instructions
with st.sidebar:
    st.image("https://img.icons8.com/color/144/null/survey.png", width=72)
    st.markdown("<h2 style='color: #1E3A8A; font-family: Outfit;'>Configuration</h2>", unsafe_allow_html=True)
    
    if is_local_mode:
        st.markdown("### 💻 Running in Local Mode")
        master_path_input = st.text_input("Target Master Database Path:", value=default_master)
    else:
        st.markdown("### ☁️ Running in Cloud Mode")
        st.info("Files will be processed in-memory. Please upload both the Raw Survey and Historical Master files.")
        master_path_input = None
        
    st.markdown("---")
    st.markdown("### How to Use")
    st.markdown("""
    1. **Upload** the newly received 2026 raw responses Excel file.
    2. *Cloud Mode*: **Upload** the Historical Master database.
    3. Click **Run Validation (Dry Run)** to check parsing and preview output datasets.
    4. Review generated charts and counts.
    5. Click **Process & Append** to prepare the final Excel file.
    6. Download the updated database!
    """)
    
    st.markdown("---")
    st.markdown("<div style='font-size: 0.8rem; color: #64748B;'>ISN Survey Automation Tool v1.1<br>© National Sports Institute of Malaysia</div>", unsafe_allow_html=True)

# Main Dashboard Title
st.markdown("<h1 class='main-title'>ISN Survey Data Integration Hub</h1>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>Streamlined cleaning, unpivoting, and appending for Sports Science & Sports Medicine satisfaction survey data</p>", unsafe_allow_html=True)

# Clean values helper
def clean_val(val, split_slash=True, age_normalize=False):
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if split_slash and "/" in val_str:
        val_str = val_str.split("/")[0].strip()
    if age_normalize:
        if "Below 18" in val_str:
            return "<18"
    return val_str

# Clean ratings helper
def clean_rating(val):
    if pd.isna(val):
        return 0
    val_str = str(val).strip()
    if val_str.upper() == "N/A":
        return 0
    if val_str and val_str[0].isdigit():
        return int(val_str[0])
    return 0

# Helper to find true max row
def get_true_max_row(ws, check_col=1):
    for r in range(ws.max_row, 0, -1):
        if ws.cell(row=r, column=check_col).value is not None:
            return r
    return 1

# Safe file pointer reset and read helpers
def load_wb_safely(source):
    if hasattr(source, "seek"):
        source.seek(0)
    return openpyxl.load_workbook(source)

def load_xls_safely(source):
    if hasattr(source, "seek"):
        source.seek(0)
    return pd.ExcelFile(source)

# Define service maps
spsc_services = [
    {"name": "Exercise Physiology", "col_received": 11, "col_ratings": [18, 25, 32, 39, 46, 53]},
    {"name": "Performance Analysis", "col_received": 12, "col_ratings": [19, 26, 33, 40, 47, 54]},
    {"name": "Biomechanics", "col_received": 13, "col_ratings": [20, 27, 34, 41, 48, 55]},
    {"name": "Sports Nutrition", "col_received": 14, "col_ratings": [21, 28, 35, 42, 49, 56]},
    {"name": "Sports Psychology", "col_received": 15, "col_ratings": [22, 29, 36, 43, 50, 57]},
    {"name": "Strength and Conditioning", "col_received": 16, "col_ratings": [23, 30, 37, 44, 51, 58]},
    {"name": "Recovery Center", "col_received": 17, "col_ratings": [24, 31, 38, 45, 52, 59]}
]

med_services = [
    {"name": "Counter services", "col": 61},
    {"name": "Sports Medicine Specialist Clinic", "col": 62},
    {"name": "Radiology services", "col": 63},
    {"name": "Sports Massage Therapy", "col": 64},
    {"name": "Medical Laboratory services", "col": 65},
    {"name": "Pharmacy services", "col": 66},
    {"name": "Physiotherapy and Rehabilitation services", "col": 67},
    {"name": "Women's Health Clinic services", "col": 68},
    {"name": "Waiting time", "col": 69},
    {"name": "Level of cleanliness and comfort of the facilities, including the waiting areas at the main counter, physiotherapy, and radiology", "col": 70}
]

q_cols = [
    "Effective communication with you",
    "Knowledgable about the sport",
    "Full commitment and involvement",
    "Benefit athletes' performance",
    "Increase coaches' and athletes' sports science knowledge",
    "Equipment Quality"
]

# Step 1: Upload raw responses
st.markdown("<div class='section-header'>Step 1: Upload Raw 2026 Responses</div>", unsafe_allow_html=True)

uploaded_raw_file = st.file_uploader("Choose the 2026 Raw Responses Excel File (.xlsx)", type=["xlsx"])

if uploaded_raw_file is not None:
    try:
        # Load Raw file
        df_raw = pd.read_excel(uploaded_raw_file)
        st.session_state.raw_preview = df_raw
        
        # Display Upload Metadata Cards
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class='metric-card'>
                <div class='card-title'>Responses Loaded</div>
                <div class='card-value'>{df_raw.shape[0]}</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class='metric-card metric-card-spsc'>
                <div class='card-title'>Raw File Columns</div>
                <div class='card-value'>{df_raw.shape[1]}</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            min_ts = df_raw.iloc[:, 0].min()
            max_ts = df_raw.iloc[:, 0].max()
            min_ts_str = pd.to_datetime(min_ts).strftime('%Y-%m-%d') if pd.notna(min_ts) else "N/A"
            max_ts_str = pd.to_datetime(max_ts).strftime('%Y-%m-%d') if pd.notna(max_ts) else "N/A"
            st.markdown(f"""
            <div class='metric-card metric-card-med'>
                <div class='card-title'>Date Range</div>
                <div style='font-size: 1.1rem; font-weight: 700; color: #1E293B; margin-top: 0.5rem;'>
                    {min_ts_str} to {max_ts_str}
                </div>
            </div>
            """, unsafe_allow_html=True)
            
        with st.expander("Preview Raw Uploaded Data"):
            st.dataframe(df_raw.head(5))
            
    except Exception as e:
        st.error(f"Error loading raw file: {e}")

else:
    # If no file uploaded and running locally, load default file
    if is_local_mode and os.path.exists(default_raw):
        st.info(f"Using default local raw responses file at: `{default_raw}`")
        if st.button("Load Default 2026 responses file"):
            try:
                df_raw = pd.read_excel(default_raw)
                st.session_state.raw_preview = df_raw
                st.rerun()
            except Exception as e:
                st.error(f"Error loading default raw file: {e}")
    else:
        st.warning("Please upload a 2026 Raw Responses file to begin.")

# Step 2: Upload or detect Master Database
st.markdown("<div class='section-header'>Step 2: Master Database Setup</div>", unsafe_allow_html=True)

uploaded_master_file = None
master_source = None

if is_local_mode:
    if os.path.exists(master_path_input):
        st.success(f"Master Database detected at: `{master_path_input}`")
        master_source = master_path_input
    else:
        st.error(f"Master database not found at `{master_path_input}`. Please correct the path in the sidebar.")
else:
    uploaded_master_file = st.file_uploader("Choose the Historical Master Database File (.xlsx)", type=["xlsx"])
    if uploaded_master_file is not None:
        st.success("Historical Master Database loaded in memory.")
        master_source = uploaded_master_file
    else:
        st.warning("Please upload the Historical Master Database File (`Survey 2022-2025 Data Source.xlsx`) to continue.")

if master_source is not None:
    try:
        xls = load_xls_safely(master_source)
        sheet_names = xls.sheet_names
        
        # Display sheet summary
        col_s1, col_s2, col_s3, col_s4, col_s5 = st.columns(5)
        sheets_info = {}
        for name in sheet_names:
            df_s = pd.read_excel(xls, name, nrows=1)
            sheets_info[name] = df_s.shape[1]
            
        with col_s1:
            st.metric("Comparison Sheet Cols", sheets_info.get("Comparison", "N/A"))
        with col_s2:
            st.metric("Form Responses 1 Cols", sheets_info.get("Form Responses 1", "N/A"))
        with col_s3:
            st.metric("Overall SpSc Cols", sheets_info.get("Overall SpSc", "N/A"))
        with col_s4:
            st.metric("Overall Sport Med Cols", sheets_info.get("Overall Sport Med", "N/A"))
        with col_s5:
            st.metric("Comparison Sport Med Cols", sheets_info.get("Comparison Sport Med", "N/A"))
            
    except Exception as e:
        st.error(f"Error analyzing master database: {e}")

# Step 3: Validate and Dry Run
st.markdown("<div class='section-header'>Step 3: Validate and Dry Run</div>", unsafe_allow_html=True)

if st.session_state.raw_preview is not None and master_source is not None:
    if st.button("Run Validation (Dry Run)", type="primary"):
        with st.spinner("Processing surveys and mapping structures..."):
            try:
                df_raw = st.session_state.raw_preview
                
                # 1. SpSc Pivoted (Overall SpSc)
                spsc_pivoted_rows = []
                spsc_unpivoted_rows = []
                
                for idx, row in df_raw.iterrows():
                    ts = row.iloc[0]
                    age = clean_val(row.iloc[2], split_slash=False)
                    gender = clean_val(row.iloc[3], split_slash=True)
                    nationality = clean_val(row.iloc[4], split_slash=True)
                    period = clean_val(row.iloc[5], split_slash=True)
                    role = clean_val(row.iloc[6], split_slash=True)
                    sport = clean_val(row.iloc[7], split_slash=True)
                    if sport.upper() == "OTHERS" and not pd.isna(row.iloc[8]):
                        sport = clean_val(row.iloc[8], split_slash=True)
                    prog = clean_val(row.iloc[9], split_slash=True)
                    comment = row.iloc[60] if not pd.isna(row.iloc[60]) else ""
                    
                    for service in spsc_services:
                        received = str(row.iloc[service["col_received"]]).strip().upper()
                        if received == "YES":
                            ratings = [clean_rating(row.iloc[col]) for col in service["col_ratings"]]
                            
                            pivoted_row = {
                                "Timestamp": ts,
                                "Age / Umur": age,
                                "Gender / Jantina": gender,
                                "Nationality / Warganegara": nationality,
                                "Period of employment / Training under program / Tempoh perkhidmatan / Latihan di bawah program": period,
                                "Role in sports setup / Peranan dalam pasukan": role,
                                "Sport / Sukan": sport,
                                "Sports Program / Program Sukan": prog,
                                "Centre Services": service["name"],
                                "Effective communication with you": ratings[0],
                                "Knowledgable about the sport": ratings[1],
                                "Full commitment and involvement": ratings[2],
                                "Benefit athletes' performance": ratings[3],
                                "Increase coaches' and athletes' sports science knowledge": ratings[4],
                                "Equipment Quality": ratings[5],
                                "Comment and feedback on services / Komen dan maklum balas perkhidmatan": comment,
                                "No / Yes": "Yes / Ya"
                            }
                            spsc_pivoted_rows.append(pivoted_row)
                            
                            # Comparison structure (Unpivoted)
                            age_comp = clean_val(row.iloc[2], split_slash=False, age_normalize=True)
                            for q_idx, q_name in enumerate(q_cols):
                                if q_idx == 0:
                                    mapped_q = "Knowledgable about the sport"
                                elif q_idx == 1:
                                    mapped_q = "Effective communication with you"
                                else:
                                    mapped_q = q_name
                                    
                                unpivoted_row = {
                                    "Timestamp": ts,
                                    "Age / Umur": age_comp,
                                    "Gender / Jantina": gender,
                                    "Nationality / Warganegara": nationality,
                                    "Period of employment / Training under program / Tempoh perkhidmatan / Latihan di bawah program": period,
                                    "Role in sports setup / Peranan dalam pasukan": role,
                                    "Sport / Sukan": sport,
                                    "Sports Program / Program Sukan": prog,
                                    "Centre Services": service["name"],
                                    "Comment and feedback on services / Komen dan maklum balas perkhidmatan": comment,
                                    "No / Yes": "Yes / Ya",
                                    "Question": mapped_q,
                                    "Rating": ratings[q_idx]
                                }
                                spsc_unpivoted_rows.append(unpivoted_row)
                                
                df_spsc_piv = pd.DataFrame(spsc_pivoted_rows)
                df_spsc_unpiv = pd.DataFrame(spsc_unpivoted_rows)
                
                # 2. Sport Med
                med_pivoted_rows = []
                med_unpivoted_rows = []
                
                for idx, row in df_raw.iterrows():
                    ts = row.iloc[0]
                    age_raw = clean_val(row.iloc[2], split_slash=False)
                    gender_raw = clean_val(row.iloc[3], split_slash=False)
                    nationality_raw = clean_val(row.iloc[4], split_slash=False)
                    period_raw = clean_val(row.iloc[5], split_slash=False)
                    role_raw = clean_val(row.iloc[6], split_slash=False)
                    sport_raw = clean_val(row.iloc[7], split_slash=False)
                    if sport_raw.upper() == "OTHERS" and not pd.isna(row.iloc[8]):
                        sport_raw = clean_val(row.iloc[8], split_slash=False)
                    prog_raw = clean_val(row.iloc[9], split_slash=False)
                    comment_med = row.iloc[71] if not pd.isna(row.iloc[71]) else ""
                    
                    ratings_med = {s["name"]: clean_rating(row.iloc[s["col"]]) for s in med_services}
                    
                    pivoted_row_med = {
                        "Timestamp": ts,
                        "Age / Umur": age_raw,
                        "Gender / Jantina": gender_raw,
                        "Nationality / Warganegara": nationality_raw,
                        "Period of employment / Training under program / Tempoh perkhidmatan / Latihan di bawah program": period_raw,
                        "Role in sports setup / Peranan dalam pasukan": role_raw,
                        "Sport / Sukan": sport_raw,
                        "Sports Program / Program Sukan": prog_raw,
                        **ratings_med,
                        "Sport Med Comment": comment_med
                    }
                    med_pivoted_rows.append(pivoted_row_med)
                    
                    # Unpivoted Sport Med
                    age_split = clean_val(row.iloc[2], split_slash=False, age_normalize=True)
                    gender_split = clean_val(row.iloc[3], split_slash=True)
                    nationality_split = clean_val(row.iloc[4], split_slash=True)
                    period_split = clean_val(row.iloc[5], split_slash=True)
                    role_split = clean_val(row.iloc[6], split_slash=True)
                    sport_split = clean_val(row.iloc[7], split_slash=True)
                    if sport_split.upper() == "OTHERS" and not pd.isna(row.iloc[8]):
                        sport_split = clean_val(row.iloc[8], split_slash=True)
                    prog_split = clean_val(row.iloc[9], split_slash=True)
                    
                    for s in med_services:
                        unpivoted_row_med = {
                            "Timestamp": ts,
                            "Age / Umur": age_split,
                            "Gender / Jantina": gender_split,
                            "Nationality / Warganegara": nationality_split,
                            "Period of employment / Training under program / Tempoh perkhidmatan / Latihan di bawah program": period_split,
                            "Role in sports setup / Peranan dalam pasukan": role_split,
                            "Sport / Sukan": sport_split,
                            "Sports Program / Program Sukan": prog_split,
                            "Services Name": s["name"],
                            "Rating": ratings_med[s["name"]],
                            "Sport Med Comment": comment_med
                        }
                        med_unpivoted_rows.append(unpivoted_row_med)
                        
                df_med_piv = pd.DataFrame(med_pivoted_rows)
                df_med_unpiv = pd.DataFrame(med_unpivoted_rows)
                
                # Save to session state
                st.session_state.spsc_piv_preview = df_spsc_piv
                st.session_state.spsc_unpiv_preview = df_spsc_unpiv
                st.session_state.med_piv_preview = df_med_piv
                st.session_state.med_unpiv_preview = df_med_unpiv
                st.session_state.dry_run_completed = True
                
                st.success("Validation and Dry Run Completed Successfully! Preview datasets below.")
                
            except Exception as e:
                st.error(f"Error during dry run validation: {e}")

# Display validation statistics and previews
if st.session_state.dry_run_completed:
    st.markdown("### Processed Datasets Statistics")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("SpSc Pivoted (Overall SpSc) Rows", len(st.session_state.spsc_piv_preview))
    with col2:
        st.metric("SpSc Unpivoted (Comparison) Rows", len(st.session_state.spsc_unpiv_preview))
    with col3:
        st.metric("Sport Med Pivoted Rows", len(st.session_state.med_piv_preview))
    with col4:
        st.metric("Sport Med Unpivoted Rows", len(st.session_state.med_unpiv_preview))
        
    # Tabs for previews
    tab1, tab2, tab3, tab4 = st.tabs([
        "Sports Science Pivoted", 
        "Sports Science Comparison", 
        "Sports Medicine Pivoted", 
        "Sports Medicine Comparison"
    ])
    
    with tab1:
        st.dataframe(st.session_state.spsc_piv_preview.head(20))
    with tab2:
        st.dataframe(st.session_state.spsc_unpiv_preview.head(20))
    with tab3:
        st.dataframe(st.session_state.med_piv_preview.head(20))
    with tab4:
        st.dataframe(st.session_state.med_unpiv_preview.head(20))
        
    # Visualizations
    st.markdown("<div class='section-header'>Data Insights & Visualizations</div>", unsafe_allow_html=True)
    col_v1, col_v2 = st.columns(2)
    
    with col_v1:
        st.markdown("#### Sports Science Responses by Service Centre")
        spsc_counts = st.session_state.spsc_piv_preview["Centre Services"].value_counts().reset_index()
        spsc_counts.columns = ["Centre Services", "Responses Count"]
        st.bar_chart(spsc_counts.set_index("Centre Services"))
        
    with col_v2:
        st.markdown("#### Sports Medicine Average Ratings by Service")
        med_avgs = []
        for s in med_services:
            avg_rating = st.session_state.med_piv_preview[s["name"]].mean()
            med_avgs.append({"Service": s["name"], "Average Rating": avg_rating})
        df_med_avgs = pd.DataFrame(med_avgs)
        st.bar_chart(df_med_avgs.set_index("Service"))

    # Step 4: Write back to master database
    st.markdown("<div class='section-header'>Step 4: Execute Write & Save</div>", unsafe_allow_html=True)
    
    if is_local_mode:
        st.warning("⚠️ Clicking the button below will modify the target Excel file in-place on your computer. A backup file will be created automatically in the same folder.")
    else:
        st.info("ℹ️ Running in cloud mode. Clicking the button below will update the workbook in memory and provide a download link.")
        
    if st.button("Process & Append to Master Database", type="secondary"):
        with st.spinner("Executing database update..."):
            try:
                # 1. Backup if local
                if is_local_mode:
                    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup_path = f"{master_path_input}.bak_{timestamp_str}"
                    shutil.copy(master_path_input, backup_path)
                    st.info(f"Created backup file: `{backup_path}`")
                
                # 2. Open workbook with openpyxl safely
                wb = load_wb_safely(master_source)
                
                # Appending Form Responses 1 (first 8 demographics columns)
                ws_fr = wb["Form Responses 1"]
                fr_headers = [cell.value for cell in ws_fr[1]]
                start_row = get_true_max_row(ws_fr) + 1
                for idx, row in st.session_state.raw_preview.iterrows():
                    row_data = {
                        "Timestamp": row.iloc[0],
                        "Age / Umur": clean_val(row.iloc[2], split_slash=False),
                        "Gender / Jantina": clean_val(row.iloc[3], split_slash=False),
                        "Nationality / Warganegara": clean_val(row.iloc[4], split_slash=False),
                        "Period of employment / Training under program / Tempoh perkhidmatan / Latihan di bawah program": clean_val(row.iloc[5], split_slash=False),
                        "Role in sports setup / Peranan dalam pasukan": clean_val(row.iloc[6], split_slash=False),
                        "Sport / Sukan": clean_val(row.iloc[7], split_slash=False),
                        "Sports Program / Program Sukan": clean_val(row.iloc[9], split_slash=False)
                    }
                    for c_idx, h in enumerate(fr_headers, 1):
                        ws_fr.cell(row=start_row, column=c_idx, value=row_data.get(h, None))
                    start_row += 1
                st.success(f"Appended {len(st.session_state.raw_preview)} rows to 'Form Responses 1'.")
                
                # Appending Overall SpSc
                ws_spsc = wb["Overall SpSc"]
                spsc_headers = [cell.value for cell in ws_spsc[1]]
                start_row = get_true_max_row(ws_spsc) + 1
                for _, row in st.session_state.spsc_piv_preview.iterrows():
                    for c_idx, h in enumerate(spsc_headers, 1):
                        val = row.get(h, None)
                        if isinstance(val, pd.Timestamp):
                            val = val.to_pydatetime()
                        ws_spsc.cell(row=start_row, column=c_idx, value=val)
                    start_row += 1
                st.success(f"Appended {len(st.session_state.spsc_piv_preview)} rows to 'Overall SpSc'.")
                
                # Appending Comparison
                ws_comp = wb["Comparison"]
                comp_headers = [cell.value for cell in ws_comp[1]]
                start_row = get_true_max_row(ws_comp) + 1
                for _, row in st.session_state.spsc_unpiv_preview.iterrows():
                    for c_idx, h in enumerate(comp_headers, 1):
                        val = row.get(h, None)
                        if isinstance(val, pd.Timestamp):
                            val = val.to_pydatetime()
                        ws_comp.cell(row=start_row, column=c_idx, value=val)
                    start_row += 1
                st.success(f"Appended {len(st.session_state.spsc_unpiv_preview)} rows to 'Comparison'.")
                
                # Appending Overall Sport Med
                ws_med = wb["Overall Sport Med"]
                
                # Insert new column for "Sports Massage Therapy" if not exists
                med_headers = [cell.value for cell in ws_med[1]]
                if "Sports Massage Therapy" not in med_headers:
                    rad_idx = None
                    for col in range(1, ws_med.max_column + 1):
                        if ws_med.cell(row=1, column=col).value == "Radiology services":
                            rad_idx = col
                            break
                    if rad_idx is not None:
                        ws_med.insert_cols(rad_idx + 1)
                        ws_med.cell(row=1, column=rad_idx + 1, value="Sports Massage Therapy")
                        med_headers = [cell.value for cell in ws_med[1]]
                        st.info("Inserted new 'Sports Massage Therapy' column into 'Overall Sport Med' sheet.")
                
                start_row = get_true_max_row(ws_med) + 1
                for _, row in st.session_state.med_piv_preview.iterrows():
                    for c_idx, h in enumerate(med_headers, 1):
                        val = row.get(h, None)
                        if isinstance(val, pd.Timestamp):
                            val = val.to_pydatetime()
                        ws_med.cell(row=start_row, column=c_idx, value=val)
                    start_row += 1
                st.success(f"Appended {len(st.session_state.med_piv_preview)} rows to 'Overall Sport Med'.")
                
                # Appending Comparison Sport Med
                ws_med_comp = wb["Comparison Sport Med"]
                med_comp_headers = [cell.value for cell in ws_med_comp[1]]
                start_row = get_true_max_row(ws_med_comp) + 1
                for _, row in st.session_state.med_unpiv_preview.iterrows():
                    for c_idx, h in enumerate(med_comp_headers, 1):
                        val = row.get(h, None)
                        if isinstance(val, pd.Timestamp):
                            val = val.to_pydatetime()
                        ws_med_comp.cell(row=start_row, column=c_idx, value=val)
                    start_row += 1
                st.success(f"Appended {len(st.session_state.med_unpiv_preview)} rows to 'Comparison Sport Med'.")
                
                # Save Workbook based on Mode
                if is_local_mode:
                    wb.save(master_path_input)
                    st.balloons()
                    st.success("Master database updated successfully in-place!")
                    
                    with open(master_path_input, "rb") as f:
                        file_bytes = f.read()
                    st.download_button(
                        label="📥 Download Updated Database Workbook",
                        data=file_bytes,
                        file_name=os.path.basename(master_path_input),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    # Save to bytes stream
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    file_bytes = buffer.getvalue()
                    
                    st.balloons()
                    st.success("Database processed and updated in memory successfully!")
                    
                    st.download_button(
                        label="📥 Download Updated Database Workbook",
                        data=file_bytes,
                        file_name="Survey_2022-2026_Updated_Data_Source.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
            except Exception as e:
                st.error(f"Error during file append execution: {e}")
